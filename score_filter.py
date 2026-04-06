import json
import pandas as pd
from pathlib import Path
from datetime import datetime
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from docx import Document

# ── Config ────────────────────────────────────────────────────────────────────

TOKEN_FILE        = "config.json"
MASTER_RESUME     = "master_resume.docx"
OUTPUT_DIR        = Path("output")
RELEVANCE_CUTOFF  = 80    # only tailor resumes with score >= this

# ── Load config ───────────────────────────────────────────────────────────────

def load_config(path: str) -> dict:
    with open(path, "r") as f:
        return json.load(f)

# ── Extract master resume text ────────────────────────────────────────────────

def extract_resume_text(docx_path: str) -> str:
    doc  = Document(docx_path)
    text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    if not text:
        raise ValueError(f"Master resume '{docx_path}' appears to be empty.")
    return text

# ── Calculate relevance score (TF-IDF cosine similarity) ─────────────────────

def calculate_relevance_score(resume_text: str, jd_text: str) -> int:
    if not jd_text or not jd_text.strip():
        return 0
    try:
        vectorizer = TfidfVectorizer(stop_words="english", ngram_range=(1, 2))
        tfidf      = vectorizer.fit_transform([resume_text, jd_text])
        score      = cosine_similarity(tfidf[0:1], tfidf[1:2])[0][0]
        return round(score * 100)
    except Exception:
        return 0

# ── Extract matched and missing keywords ──────────────────────────────────────

def extract_keywords(resume_text: str, jd_text: str, top_n: int = 15) -> tuple[list, list]:
    if not jd_text or not jd_text.strip():
        return [], []
    try:
        vectorizer = TfidfVectorizer(stop_words="english", max_features=50, ngram_range=(1, 2))
        vectorizer.fit([jd_text])
        jd_keywords  = vectorizer.get_feature_names_out().tolist()
        resume_lower = resume_text.lower()
        matched      = [kw for kw in jd_keywords if kw.lower() in resume_lower]
        missing      = [kw for kw in jd_keywords if kw.lower() not in resume_lower]
        return matched[:top_n], missing[:top_n]
    except Exception:
        return [], []

# ── Find most recent Excel in output folder ───────────────────────────────────

def find_latest_excel(output_dir: Path) -> Path:
    folders = sorted(output_dir.glob("????-??-??"), reverse=True)
    if not folders:
        raise FileNotFoundError("No daily output folders found. Run scraper.py first.")
    for folder in folders:
        excels = sorted(folder.glob("*.xlsx"), reverse=True)
        # Skip the report file
        excels = [e for e in excels if "report" not in e.name]
        if excels:
            return excels[0]
    raise FileNotFoundError("No Excel job files found. Run scraper.py first.")

# ── Score all jobs and filter ─────────────────────────────────────────────────

def score_and_filter(excel_path: Path, resume_text: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    df   = pd.read_excel(excel_path)
    rows = []

    print(f"\n📊 Scoring {len(df)} jobs against master resume...")
    print(f"   Cutoff: Relevance Score >= {RELEVANCE_CUTOFF}\n")

    for _, job in df.iterrows():
        company = str(job.get("Company")         or "").strip()
        title   = str(job.get("Job Title")       or "").strip()
        jd      = str(job.get("Job Description") or "").strip()

        if not jd or jd.lower() in ("nan", "none", ""):
            score   = 0
            matched = []
            missing = []
        else:
            score           = calculate_relevance_score(resume_text, jd)
            matched, missing = extract_keywords(resume_text, jd)

        flag = "✅ PASS" if score >= RELEVANCE_CUTOFF else "❌ SKIP"
        print(f"  {flag}  {score:>3}/100  {company} — {title}")

        rows.append({
            "Company":          company,
            "Job Title":        title,
            "Location":         job.get("Location"),
            "Job URL":          job.get("Job URL"),
            "Relevance Score":  score,
            "Matched Keywords": ", ".join(matched),
            "Missing Keywords": ", ".join(missing),
            "Job Description":  jd,
        })

    scored_df = pd.DataFrame(rows)
    scored_df.sort_values(
        by=["Relevance Score", "Company"],
        ascending=[False, True],
        inplace=True,
        na_position="last",
    )

    passed_df = scored_df[scored_df["Relevance Score"] >= RELEVANCE_CUTOFF].copy()
    skipped_df = scored_df[scored_df["Relevance Score"] < RELEVANCE_CUTOFF].copy()

    print(f"\n{'─'*50}")
    print(f"  ✅ Passed  ({'>='}{RELEVANCE_CUTOFF}): {len(passed_df)} jobs")
    print(f"  ❌ Skipped ({'< '}{RELEVANCE_CUTOFF}): {len(skipped_df)} jobs")
    print(f"{'─'*50}")

    return passed_df, scored_df

# ── Save scored Excel ─────────────────────────────────────────────────────────

def save_scored_excel(scored_df: pd.DataFrame, passed_df: pd.DataFrame, daily_dir: Path) -> Path:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils  import get_column_letter

    timestamp  = datetime.now().strftime("%H%M%S")
    out_path   = daily_dir / f"scored_jobs_{timestamp}.xlsx"

    green_fill  = PatternFill("solid", fgColor="E2EFDA")
    red_fill    = PatternFill("solid", fgColor="FFDDC1")
    header_fill = PatternFill("solid", fgColor="1F497D")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin        = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    display_cols = ["Company", "Job Title", "Location", "Job URL",
                    "Relevance Score", "Matched Keywords", "Missing Keywords"]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:

        # ── Sheet 1: All jobs scored ──────────────────────────────────────────
        scored_df[display_cols].to_excel(writer, index=False, sheet_name="All Jobs")
        ws = writer.sheets["All Jobs"]

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin

        for row in ws.iter_rows(min_row=2):
            score = row[4].value or 0   # Relevance Score col
            fill  = green_fill if score >= RELEVANCE_CUTOFF else red_fill
            for cell in row:
                cell.fill      = fill
                cell.border    = thin
                cell.font      = Font(size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            # Bold + color the score cell
            row[4].font      = Font(
                bold=True, size=10,
                color="375623" if score >= RELEVANCE_CUTOFF else "9C0006"
            )
            row[4].alignment = Alignment(horizontal="center", vertical="top")

        col_widths = {"Company": 25, "Job Title": 35, "Location": 20,
                      "Job URL": 40, "Relevance Score": 12,
                      "Matched Keywords": 40, "Missing Keywords": 40}
        for i, col in enumerate(display_cols, 1):
            ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col, 20)
        ws.freeze_panes = "A2"

        # ── Sheet 2: Passed jobs only (for resume tailor) ─────────────────────
        passed_df[display_cols].to_excel(writer, index=False, sheet_name="Passed (≥80)")
        ws2 = writer.sheets["Passed (≥80)"]

        for cell in ws2[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin

        for row in ws2.iter_rows(min_row=2):
            for cell in row:
                cell.fill      = green_fill
                cell.border    = thin
                cell.font      = Font(size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            row[4].font      = Font(bold=True, size=10, color="375623")
            row[4].alignment = Alignment(horizontal="center", vertical="top")

        for i, col in enumerate(display_cols, 1):
            ws2.column_dimensions[get_column_letter(i)].width = col_widths.get(col, 20)
        ws2.freeze_panes = "A2"

    print(f"\n💾 Scored Excel saved → {out_path.resolve()}")
    return out_path

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    import sys

    # Accept optional Excel path as argument, else auto-detect latest
    if len(sys.argv) >= 2:
        excel_path = Path(sys.argv[1])
        if not excel_path.exists():
            print(f"❌ File not found: {excel_path}")
            sys.exit(1)
    else:
        excel_path = find_latest_excel(OUTPUT_DIR)

    daily_dir   = excel_path.parent
    resumes_dir = daily_dir / "resumes"
    resumes_dir.mkdir(exist_ok=True)

    print(f"📄 Jobs file   : {excel_path}")
    print(f"📄 Resume file : {MASTER_RESUME}")

    resume_text            = extract_resume_text(MASTER_RESUME)
    passed_df, scored_df   = score_and_filter(excel_path, resume_text)
    scored_excel           = save_scored_excel(scored_df, passed_df, daily_dir)

    if passed_df.empty:
        print(f"\n⚠️  No jobs passed the {RELEVANCE_CUTOFF} threshold. Try lowering RELEVANCE_CUTOFF.")
        return

    # Save passed jobs as a separate Excel for resume_tailor
    passed_path = daily_dir / "passed_jobs.xlsx"
    # Include Job Description for resume_tailor to use
    passed_df_full = passed_df.copy()
    passed_df_full.to_excel(passed_path, index=False)
    print(f"📋 Passed jobs  → {passed_path.resolve()}")

    run_tailor = input(
        f"\n🤖 Run resume tailor on {len(passed_df)} passed job(s)? (y/n): "
    ).strip().lower()

    if run_tailor == "y":
        from resume_tailor import tailor_resumes
        tailor_resumes(passed_path, resumes_dir)

if __name__ == "__main__":
    main()