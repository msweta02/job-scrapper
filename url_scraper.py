import requests
import json
import re
import pandas as pd
from pathlib import Path
from datetime import datetime
from bs4 import BeautifulSoup
from urllib.parse import urlparse

# ── Config ────────────────────────────────────────────────────────────────────

TOKEN_FILE = "config.json"
URLS_FILE  = "job_urls.txt"
OUTPUT_DIR = Path("output")

MASTER_RESUME = "master_resume.docx"
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

# ── Load URLs ─────────────────────────────────────────────────────────────────

def load_urls(path: str) -> list[str]:
    file = Path(path)
    if not file.exists():
        raise FileNotFoundError(f"URL file not found: {path}")
    urls = []
    for line in file.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line and not line.startswith("#"):
            urls.append(line)
    if not urls:
        raise ValueError(f"No URLs found in {path}")
    print(f"📋 Loaded {len(urls)} URL(s) from {path}")
    return urls

# ── Create daily output folders ───────────────────────────────────────────────

def get_daily_dirs() -> tuple[Path, Path]:
    today       = datetime.now().strftime("%Y-%m-%d")
    daily_dir   = OUTPUT_DIR / today
    resumes_dir = daily_dir / "resumes"
    daily_dir.mkdir(parents=True, exist_ok=True)
    resumes_dir.mkdir(parents=True, exist_ok=True)
    print(f"📁 Output folder: {daily_dir.resolve()}")
    return daily_dir, resumes_dir

# ── Detect portal type ────────────────────────────────────────────────────────

def detect_portal(url: str) -> str:
    domain = urlparse(url).netloc.lower()
    if "linkedin.com"    in domain: return "linkedin"
    if "greenhouse.io"   in domain: return "greenhouse"
    if "lever.co"        in domain: return "lever"
    if "workday.com"     in domain: return "workday"
    if "myworkdayjobs"   in domain: return "workday"
    if "smartrecruiters" in domain: return "smartrecruiters"
    if "icims.com"       in domain: return "icims"
    if "jobvite.com"     in domain: return "jobvite"
    if "indeed.com"      in domain: return "indeed"
    return "generic"

# ── Greenhouse — use their free API ──────────────────────────────────────────

def scrape_greenhouse(url: str) -> dict:
    m = re.search(r"greenhouse\.io/([^/]+)/jobs/(\d+)", url)
    if m:
        board, job_id = m.group(1), m.group(2)
        api = f"https://boards-api.greenhouse.io/v1/boards/{board}/jobs/{job_id}"
        res = requests.get(api, timeout=15)
        if res.ok:
            d = res.json()
            return {
                "Company":         board.replace("-", " ").title(),
                "Job Title":       d.get("title"),
                "Location":        d.get("location", {}).get("name"),
                "Job Description": BeautifulSoup(d.get("content", ""), "html.parser").get_text("\n"),
                "Job URL":         url,
                "Portal":          "Greenhouse",
            }
    return scrape_generic(url)

# ── Lever — use their free API ────────────────────────────────────────────────

def scrape_lever(url: str) -> dict:
    m = re.search(r"lever\.co/([^/]+)/([^/?]+)", url)
    if m:
        company, job_id = m.group(1), m.group(2)
        api = f"https://api.lever.co/v0/postings/{company}/{job_id}"
        res = requests.get(api, timeout=15)
        if res.ok:
            d    = res.json()
            desc = BeautifulSoup(
                d.get("descriptionPlain", "") + "\n" + d.get("additionalPlain", ""),
                "html.parser"
            ).get_text("\n")
            return {
                "Company":         company.replace("-", " ").title(),
                "Job Title":       d.get("text"),
                "Location":        d.get("categories", {}).get("location"),
                "Job Description": desc.strip(),
                "Job URL":         url,
                "Portal":          "Lever",
            }
    return scrape_generic(url)

# ── Generic scraper (Workday, iCIMS, company careers pages, etc.) ─────────────

def scrape_generic(url: str) -> dict:
    portal  = detect_portal(url).title()
    domain  = urlparse(url).netloc.replace("www.", "").split(".")[0].replace("-", " ").title()

    try:
        res = requests.get(url, headers=HEADERS, timeout=20)
        res.raise_for_status()
        soup = BeautifulSoup(res.text, "html.parser")

        # Remove noise tags
        for tag in soup(["script", "style", "nav", "footer", "header"]):
            tag.decompose()

        # ── Title ─────────────────────────────────────────────────────────────
        title = None
        for sel in [
            "h1.job-title", "h1.posting-headline", "h1[data-ui='job-title']",
            "h1.jobTitle", ".job-title h1", "h1"
        ]:
            el = soup.select_one(sel)
            if el and el.get_text(strip=True):
                title = el.get_text(strip=True)
                break
        if not title:
            og = soup.find("meta", property="og:title")
            title = og["content"] if og else soup.title.string if soup.title else None

        # ── Company ───────────────────────────────────────────────────────────
        company = None
        for sel in [".company-name", ".employer-name", "[data-company]"]:
            el = soup.select_one(sel)
            if el:
                company = el.get_text(strip=True)
                break
        if not company:
            og = soup.find("meta", property="og:site_name")
            company = og["content"] if og else domain

        # ── Location ──────────────────────────────────────────────────────────
        location = None
        for sel in [
            ".location", ".job-location", "[data-ui='job-location']",
            "[itemprop='jobLocation']", ".posting-categories .sort-by-time",
            ".jobs-unified-top-card__bullet", ".location-text"
        ]:
            el = soup.select_one(sel)
            if el and el.get_text(strip=True):
                location = el.get_text(strip=True)
                break

        # ── Description ───────────────────────────────────────────────────────
        description = None
        for sel in [
            ".job-description", "#job-description", "[data-ui='job-description']",
            ".description__text", ".jobDescriptionContent", ".posting-description",
            ".careers-job-description", "article", "main", "#main-content"
        ]:
            el = soup.select_one(sel)
            if el and len(el.get_text(strip=True)) > 200:
                description = el.get_text("\n", strip=True)
                break
        if not description:
            description = soup.get_text("\n", strip=True)

        # Trim excess whitespace
        description = re.sub(r"\n{3,}", "\n\n", description).strip()

        return {
            "Company":         company,
            "Job Title":       title,
            "Location":        location,
            "Job Description": description,
            "Job URL":         url,
            "Portal":          portal,
        }

    except Exception as e:
        print(f"         ⚠️  Error scraping {url}: {e}")
        return {
            "Company":         domain,
            "Job Title":       None,
            "Location":        None,
            "Job Description": None,
            "Job URL":         url,
            "Portal":          portal,
            "Error":           str(e),
        }

# ── Route each URL to the right scraper ──────────────────────────────────────

def scrape_url(url: str) -> dict:
    portal = detect_portal(url)
    if portal == "greenhouse":
        return scrape_greenhouse(url)
    elif portal == "lever":
        return scrape_lever(url)
    else:
        return scrape_generic(url)

# ── Save to Excel ─────────────────────────────────────────────────────────────

def save_to_excel(jobs: list[dict], daily_dir: Path) -> Path:
    if not jobs:
        print("⚠️  No jobs to save.")
        return None

    # Ensure all rows have the same columns
    cols = ["Company", "Job Title", "Location", "Job URL",
            "Portal", "Job Description", "Error"]
    for j in jobs:
        for c in cols:
            j.setdefault(c, None)

    df = pd.DataFrame(jobs)[cols]
    df.sort_values(by="Company", ascending=True, inplace=True, na_position="last")

    timestamp = datetime.now().strftime("%H%M%S")
    out_path  = daily_dir / f"url_jobs_{timestamp}.xlsx"

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Jobs")
        ws = writer.sheets["Jobs"]

        from openpyxl.styles import Alignment, PatternFill, Font
        header_fill = PatternFill("solid", fgColor="1F497D")
        header_font = Font(color="FFFFFF", bold=True, size=10)

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.columns:
            max_len = max(
                max((len(line) for line in str(cell.value).split("\n")), default=0)
                if cell.value else 0
                for cell in col
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

        # Wrap text + top align all cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.freeze_panes = "A2"

    print(f"💾 Saved to: {out_path.resolve()}")
    return out_path

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    urls                   = load_urls(URLS_FILE)
    daily_dir, resumes_dir = get_daily_dirs()

    jobs = []
    for i, url in enumerate(urls, 1):
        print(f"  [{i}/{len(urls)}] 🌐 {url}")
        job = scrape_url(url)
        print(f"         ✓ {job.get('Company')} — {job.get('Job Title')} ({job.get('Portal')})")
        jobs.append(job)

    excel_path = save_to_excel(jobs, daily_dir)

    if excel_path:
        from score_filter import score_and_filter, save_scored_excel, extract_resume_text
        resume_text          = extract_resume_text(MASTER_RESUME)
        passed_df, scored_df = score_and_filter(excel_path, resume_text)
        save_scored_excel(scored_df, passed_df, daily_dir)

        if not passed_df.empty:
            passed_path = daily_dir / "passed_jobs.xlsx"
            passed_df.to_excel(passed_path, index=False)
            print(f"📋 Passed jobs → {passed_path.resolve()}")

            run_tailor = input(
                f"\n🤖 Run resume tailor on {len(passed_df)} passed job(s)? (y/n): "
            ).strip().lower()
            if run_tailor == "y":
                from resume_tailor import tailor_resumes
                tailor_resumes(passed_path, resumes_dir)
        else:
            print(f"\n⚠️  No jobs passed the relevance threshold.")

if __name__ == "__main__":
    main()